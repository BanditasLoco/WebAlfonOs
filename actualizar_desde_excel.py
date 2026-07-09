"""
Actualiza el Monitor de inundaciones a partir de un parte en Excel.

Flujo:
  1. Llenas 'parte_diario_template.xlsx' (hoja 'Parte') con las novedades del día.
  2. Corres:  python actualizar_desde_excel.py [parte.xlsx] [base.json] [salida.json]
  3. Importas la salida en el dashboard.

Solo parchea los cantones listados; el resto conserva su estado.
Cantón existente -> actualiza estado/afectados/observación/tipo/ruta/conectividad/fecha.
Cantón nuevo (con Lat, Lng y Nivel) -> se agrega. Idempotente: no duplica.
Las celdas vacías de Tipo/Ruta/Conectividad NO borran el dato anterior.
"""
import json, sys, datetime
from openpyxl import load_workbook

PARTE  = sys.argv[1] if len(sys.argv) > 1 else "parte_diario_template.xlsx"
BASE   = sys.argv[2] if len(sys.argv) > 2 else "monitor_base.json"
SALIDA = sys.argv[3] if len(sys.argv) > 3 else "monitor_ACTUALIZADO.json"
FECHA  = datetime.date.today().isoformat()
SCORE  = {"CRÍTICO": 4, "ALTO": 3, "MEDIO-ALTO": 2, "MEDIO": 1}
ESTADOS = {"Normal", "Amarilla", "Naranja", "Roja"}

# Orden de columnas en la hoja 'Parte'
COLS = ["provincia", "canton", "estado", "posibles_afectadas", "consultoras", "consultoras_base", "obs",
        "tipo", "rutas", "conectividad", "lat", "lng", "nivel"]


def leer_parte(path):
    ws = load_workbook(path, data_only=True)["Parte"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(COLS, row))
        if not d.get("canton"):
            continue
        out.append(d)
    return out


def txt(v):
    return ("" if v is None else str(v)).strip()


def main():
    recs = json.load(open(BASE, encoding="utf-8"))
    for r in recs:                       # asegurar campos nuevos
        r.setdefault("tipo", ""); r.setdefault("rutas", ""); r.setdefault("conectividad", ""); r.setdefault("consultoras_afectadas", 0); r.setdefault("consultoras_base", 0); r.setdefault("posibles_afectadas", r.pop("afectados",0) if "afectados" in r else r.get("posibles_afectadas",0))
    idx = {r["canton"]: r for r in recs}
    sig = max((int(str(r["id"])[1:]) for r in recs if str(r["id"])[1:].isdigit()), default=0) + 1

    parchea = nuevo = avisos = 0
    for d in leer_parte(PARTE):
        canton = txt(d["canton"])
        estado = txt(d.get("estado")) or "Normal"
        if estado not in ESTADOS:
            print(f"  ! '{canton}': estado '{estado}' invalido, se omite"); avisos += 1; continue
        pos = int(d.get("posibles_afectadas") or 0)
        con = int(d.get("consultoras") or 0)
        cbase = int(d.get("consultoras_base") or 0)

        if canton in idx:                                  # PARCHE
            r = idx[canton]
            r.update(estado=estado, posibles_afectadas=pos, consultoras_afectadas=con, fecha=FECHA)
            if cbase: r["consultoras_base"]=cbase
            for campo in ("obs", "tipo", "rutas", "conectividad"):  # solo si viene dato
                if txt(d.get(campo)):
                    r[campo] = txt(d[campo])
            parchea += 1
        elif d.get("lat") not in (None, "") and d.get("lng") not in (None, ""):  # ALTA
            nivel = txt(d.get("nivel")) or "ALTO"
            r = {"id": f"c{sig}", "provincia": txt(d.get("provincia")) or "?",
                 "canton": canton, "nivel": nivel, "score": SCORE.get(nivel, 3),
                 "pct": "n/d", "cuenca": False, "lat": float(d["lat"]), "lng": float(d["lng"]),
                 "estado": estado, "posibles_afectadas": pos, "consultoras_afectadas": con, "consultoras_base": cbase, "fecha": FECHA,
                 "obs": txt(d.get("obs")), "tipo": txt(d.get("tipo")),
                 "rutas": txt(d.get("rutas")), "conectividad": txt(d.get("conectividad"))}
            recs.append(r); idx[canton] = r; sig += 1; nuevo += 1
        else:
            print(f"  ! '{canton}': no existe y faltan Lat/Lng/Nivel para crearlo"); avisos += 1

    json.dump(recs, open(SALIDA, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    activos = [r for r in recs if r["estado"] != "Normal"]
    total = sum(r.get("posibles_afectadas",0) for r in recs)
    por = {e: sum(1 for r in recs if r["estado"] == e) for e in ["Roja", "Naranja", "Amarilla", "Normal"]}
    print(f"Parte    : {PARTE}")
    print(f"Resultado: {parchea} parcheados, {nuevo} nuevos, {avisos} avisos")
    print(f"Salida   : {SALIDA}  ({len(recs)} cantones)")
    print(f"Activos  : {len(activos)} con alerta, {total:,} posibles afectadas".replace(",", "."))
    print(f"Por estado: {por}")


if __name__ == "__main__":
    main()
